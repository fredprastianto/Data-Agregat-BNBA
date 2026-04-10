import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  KONFIGURASI — Sesuaikan jika perlu
# ─────────────────────────────────────────────

INPUT_FILE  = "BNBA Jadup Tapanuli Tengah.xlsx"
SHEET_NAME  = "bnba"           # <-- nama sheet yang berisi data individu
OUTPUT_FILE = "agregat_per_desa.xlsx"

GROUP_BY_COLS = ["desa", "kecamatan", "kab"]
EXCLUDE_COLS  = ["No", "nama", "no_kk", "no_kk_dtsen"]

COLOR_HEADER_BG   = "1F4E79"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_ALT_ROW     = "EBF3FB"
COLOR_TOTAL_BG    = "1F4E79"

# ─────────────────────────────────────────────


def baca_excel(path: str, sheet: str) -> pd.DataFrame:
    try:
        df = pd.read_excel(path, sheet_name=sheet)
        print(f"[OK] File dibaca : {path}")
        print(f"     Sheet       : {sheet}")
        print(f"     Jumlah baris: {len(df):,}")
        print(f"     Jumlah kolom: {len(df.columns)}")
        return df
    except FileNotFoundError:
        sys.exit(f"[ERROR] File tidak ditemukan: {path}\n"
                 f"        Pastikan file ada di folder yang sama dengan script ini.")
    except Exception as e:
        sys.exit(f"[ERROR] Gagal membaca file: {e}")


def agregasi(df: pd.DataFrame) -> pd.DataFrame:
    for col in GROUP_BY_COLS:
        if col not in df.columns:
            sys.exit(
                f"[ERROR] Kolom '{col}' tidak ditemukan.\n"
                f"        Kolom tersedia: {df.columns.tolist()}\n"
                f"        Sesuaikan GROUP_BY_COLS di bagian KONFIGURASI."
            )

    semua_exclude = set(EXCLUDE_COLS) | set(GROUP_BY_COLS)
    value_cols = [
        c for c in df.columns
        if c not in semua_exclude and pd.api.types.is_numeric_dtype(df[c])
    ]

    print(f"\n[INFO] Kolom yang di-sum ({len(value_cols)} kolom):")
    for c in value_cols:
        print(f"       - {c}")

    agg    = df.groupby(GROUP_BY_COLS)[value_cols].sum().reset_index()
    jumlah = df.groupby(GROUP_BY_COLS[0]).size().reset_index(name="jumlah_penerima")
    agg    = agg.merge(jumlah, on=GROUP_BY_COLS[0])

    return agg[GROUP_BY_COLS + ["jumlah_penerima"] + value_cols]


def tulis_excel(df: pd.DataFrame, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Agregat per Desa"

    def thin_border():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    def hdr_style(cell):
        cell.font      = Font(name="Arial", bold=True, color=COLOR_HEADER_TEXT, size=10)
        cell.fill      = PatternFill("solid", start_color=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()

    def data_style(cell, bold=False, number_fmt=None, align="center", bg=None):
        fg = COLOR_HEADER_TEXT if bg == COLOR_TOTAL_BG else "000000"
        cell.font      = Font(name="Arial", bold=bold, size=10, color=fg)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = thin_border()
        if number_fmt:
            cell.number_format = number_fmt
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)

    headers = list(df.columns)
    n_cols  = len(headers)

    # Baris 1: Judul
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws["A1"]
    t.value     = "DATA AGREGAT PER DESA — BNBA JADUP TAPANULI TENGAH"
    t.font      = Font(name="Arial", bold=True, size=13, color=COLOR_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 4

    # Baris 3: Header kolom
    LABEL = {
        "desa"                                : "Desa",
        "kecamatan"                           : "Kecamatan",
        "kab"                                 : "Kabupaten",
        "jumlah_penerima"                     : "Jumlah\nPenerima",
        "rusak_ringan"                        : "Rusak\nRingan",
        "rusak_sedang"                        : "Rusak\nSedang",
        "rusak_berat"                         : "Rusak\nBerat",
        "rusak_hanyut"                        : "Rusak\nHanyut",
        "rusak_lainnya"                       : "Rusak\nLainnya",
        "rusak_tidak_ada_informasi"           : "Rusak\nTdk Ada Info",
        "desil_nasional"                      : "Desil\nNasional",
        "status_kepemilikan_rumah"            : "Status\nKepemilikan",
        "jenis_lantai_terluas"                : "Jenis\nLantai",
        "luas_lantai"                         : "Luas\nLantai",
        "jenis_dinding_terluas"               : "Jenis\nDinding",
        "jenis_atap_terluas"                  : "Jenis\nAtap",
        "sumber_air_minum_utama"              : "Sumber\nAir Minum",
        "sumber_penerangan_utama"             : "Sumber\nPenerangan",
        "bahan_bakar_utama_memasak"           : "Bahan\nBakar",
        "fasilitas_bab"                       : "Fasilitas\nBAB",
        "jenis_kloset"                        : "Jenis\nKloset",
        "pembuangan_akhir_tinja"              : "Pembuangan\nTinja",
        "kepemilikan_aset"                    : "Kepemilikan\nAset",
        "aset_bergerak_tabung_gas"            : "Tabung\nGas",
        "aset_bergerak_lemari_es"             : "Lemari\nEs",
        "aset_bergerak_ac"                    : "AC",
        "aset_bergerak_pemanas_air"           : "Pemanas\nAir",
        "aset_bergerak_telepon_rumah"         : "Telepon\nRumah",
        "aset_bergerak_tv_datar"              : "TV\nDatar",
        "aset_bergerak_emas_perhiasan"        : "Emas /\nPerhiasan",
        "aset_bergerak_komputer_laptop_tablet": "Komputer /\nLaptop",
        "aset_bergerak_sepeda_motor"          : "Sepeda\nMotor",
        "aset_bergerak_sepeda"                : "Sepeda",
        "aset_bergerak_mobil"                 : "Mobil",
        "aset_bergerak_perahu"                : "Perahu",
        "aset_bergerak_kapal_perahu_motor"    : "Kapal\nMotor",
        "aset_bergerak_smartphone"            : "Smart-\nphone",
        "aset_tidak_bergerak_lahan_lainnya"   : "Lahan\nLainnya",
        "aset_tidak_bergerak_rumah_lainnya"   : "Rumah\nLainnya",
        "Total Nilai Bantuan Jadup"           : "Total Nilai\nBantuan Jadup",
    }

    for c_idx, col in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c_idx, value=LABEL.get(col, col))
        hdr_style(cell)
    ws.row_dimensions[3].height = 45

    # Baris data
    DATA_START_ROW = 4
    alt_fill = PatternFill("solid", start_color=COLOR_ALT_ROW)

    for r_idx, row in df.iterrows():
        excel_row = r_idx + DATA_START_ROW
        use_alt   = (r_idx % 2 == 1)

        for c_idx, col in enumerate(headers, 1):
            val  = row[col]
            val  = 0 if pd.isna(val) else val
            cell = ws.cell(row=excel_row, column=c_idx, value=val)

            if col in ["desa", "kecamatan", "kab"]:
                data_style(cell, align="left")
            elif col == "jumlah_penerima":
                data_style(cell, align="center")
            else:
                data_style(cell, number_fmt="#,##0", align="right")

            if use_alt:
                cell.fill = alt_fill

    # Baris TOTAL
    total_row = DATA_START_ROW + len(df)
    ws.merge_cells(
        start_row=total_row, start_column=1,
        end_row=total_row,   end_column=len(GROUP_BY_COLS)
    )
    lbl = ws.cell(row=total_row, column=1, value="TOTAL")
    data_style(lbl, bold=True, align="center", bg=COLOR_TOTAL_BG)
    for ci in range(2, len(GROUP_BY_COLS) + 1):
        data_style(ws.cell(row=total_row, column=ci), bold=True, bg=COLOR_TOTAL_BG)

    for c_idx, col in enumerate(headers, 1):
        if c_idx <= len(GROUP_BY_COLS):
            continue
        col_letter = get_column_letter(c_idx)
        formula = f"=SUM({col_letter}{DATA_START_ROW}:{col_letter}{total_row - 1})"
        cell = ws.cell(row=total_row, column=c_idx, value=formula)
        data_style(cell, bold=True, number_fmt="#,##0", align="right", bg=COLOR_TOTAL_BG)

    # Lebar kolom
    LEBAR = {
        "desa": 22, "kecamatan": 18, "kab": 18,
        "jumlah_penerima": 10, "Total Nilai Bantuan Jadup": 18,
    }
    for c_idx, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = LEBAR.get(col, 9)

    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=len(GROUP_BY_COLS) + 1)

    wb.save(output_path)
    print(f"\n[OK] File tersimpan : {output_path}")
    print(f"     Jumlah desa     : {len(df)}")
    print(f"     Kolom output    : {len(headers)}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 55)
    print("  AGREGASI DATA PER DESA")
    print("=" * 55)

    df_raw = baca_excel(INPUT_FILE, SHEET_NAME)
    df_agg = agregasi(df_raw)
    tulis_excel(df_agg, OUTPUT_FILE)

    print("\nSelesai! Buka file output untuk melihat hasilnya.")
    print("=" * 55)